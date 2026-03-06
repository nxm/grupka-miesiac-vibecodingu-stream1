import io
import httpx
import openpyxl
import cache

NBP_URL = "https://static.nbp.pl/dane/rynek-nieruchomosci/ceny_mieszkan.xlsx"

CITIES = [
    "Białystok", "Bydgoszcz", "Gdańsk", "Gdynia", "Katowice",
    "Kielce", "Kraków", "Lublin", "Łódź", "Olsztyn",
    "Opole", "Poznań", "Rzeszów", "Szczecin", "Warszawa",
    "Wrocław", "Zielona Góra",
]

# Sheets: "Rynek pierwotny" and "Rynek wtórny"
# Each sheet has two column groups: offer prices (cols 1-17) and transaction prices (cols 24-40)
# Row 7 (0-indexed=6) has city names as headers, data starts from row 8 (0-indexed=7)

OFFER_COLS = list(range(1, 18))       # cols B-R (17 cities)
TRANSACTION_COLS = list(range(24, 41))  # cols Y-AO (17 cities)


def _match_city(raw: str) -> str | None:
    """Match a raw header string to a canonical city name."""
    clean = raw.strip().rstrip("*")
    for city in CITIES:
        if city.lower() in clean.lower() or clean.lower() in city.lower():
            return city
    return None


async def _download_xlsx() -> bytes:
    cache_key = "nbp_xlsx_raw"
    cached = cache.get(cache_key)
    if cached:
        return bytes.fromhex(cached)

    async with httpx.AsyncClient(timeout=60, follow_redirects=True) as client:
        resp = await client.get(NBP_URL)
        resp.raise_for_status()
        raw = resp.content

    cache.set(cache_key, raw.hex())
    return raw


def _parse_sheet(ws, market: str) -> dict[str, list[dict]]:
    """Parse a sheet into offer and transaction records."""
    rows = list(ws.iter_rows(values_only=True))

    # Find header row with "Kwartał"
    header_idx = None
    for i, row in enumerate(rows):
        if row and row[0] and str(row[0]).strip().lower() == "kwartał":
            header_idx = i
            break

    if header_idx is None:
        return {}

    header = rows[header_idx]

    # Map column index -> city name for offer section
    offer_cities = {}
    for col_idx in OFFER_COLS:
        if col_idx < len(header) and header[col_idx]:
            city = _match_city(str(header[col_idx]))
            if city:
                offer_cities[col_idx] = city

    # Map column index -> city name for transaction section
    transaction_cities = {}
    for col_idx in TRANSACTION_COLS:
        if col_idx < len(header) and header[col_idx]:
            city = _match_city(str(header[col_idx]))
            if city:
                transaction_cities[col_idx] = city

    offer_records = []
    transaction_records = []

    for row in rows[header_idx + 1:]:
        if not row or not row[0]:
            continue

        quarter = str(row[0]).strip()
        # Quarter format: "III 2006", "IV 2006", "I 2007", etc.
        if not any(quarter.startswith(prefix) for prefix in ("I ", "II ", "III ", "IV ")):
            continue

        # Parse offer prices
        for col_idx, city in offer_cities.items():
            if col_idx < len(row) and row[col_idx] is not None:
                try:
                    price = float(row[col_idx])
                    offer_records.append({
                        "city": city,
                        "quarter": quarter,
                        "price": round(price, 2),
                    })
                except (ValueError, TypeError):
                    continue

        # Parse transaction prices
        for col_idx, city in transaction_cities.items():
            if col_idx < len(row) and row[col_idx] is not None:
                try:
                    price = float(row[col_idx])
                    transaction_records.append({
                        "city": city,
                        "quarter": quarter,
                        "price": round(price, 2),
                    })
                except (ValueError, TypeError):
                    continue

    return {
        f"{market}_offer": offer_records,
        f"{market}_transaction": transaction_records,
    }


async def get_prices() -> dict:
    """Download and parse NBP price data, return by (market, type)."""
    cache_key = "nbp_prices_parsed_v2"
    cached = cache.get(cache_key)
    if cached:
        return cached

    xlsx_bytes = await _download_xlsx()
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)

    result = {}
    for name in wb.sheetnames:
        lower = name.lower()
        if "pierwotny" in lower:
            result.update(_parse_sheet(wb[name], "primary"))
        elif "wtórn" in lower:
            result.update(_parse_sheet(wb[name], "secondary"))

    cache.set(cache_key, result)
    return result


async def get_prices_for_city(city: str) -> dict:
    """Get all price data for a specific city."""
    all_prices = await get_prices()
    city_data = {}
    for key, records in all_prices.items():
        city_data[key] = [r for r in records if r["city"].lower() == city.lower()]
    return city_data
